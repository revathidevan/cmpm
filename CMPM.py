import pandas as pd
import os
import openpyxl
from datetime import datetime
import win32com.client as win32
import time

def modify_excel_file():
    excel = None
    wb = None
    try:
        # Get the directory where the script is located
        script_dir = os.path.dirname(os.path.abspath(__file__))
        input_file = os.path.join(script_dir, 'cmpm.xlsx')
        
        # Read the entire Excel file
        df_full = pd.read_excel(input_file, header=None)
        
        # Keep first 8 rows exactly as they are (rows 1-7 and empty row 8)
        first_rows = df_full.iloc[:8].copy()
        
        # Get the header row (9th row)
        header_row = df_full.iloc[8].tolist()  # Convert to list
        
        # Get the data rows (after 9th row)
        data_rows = df_full.iloc[9:].copy()
        data_rows.columns = header_row  # Set the 9th row as header for data
        
        # Filter Cost Center to EY1Z18000
        filtered_data = data_rows[data_rows['Cost Center'] == 'EY1Z18000']
        
        # Get list of month columns
        month_columns = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                        'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        
        # Show available months and ask user
        print("\nAvailable months:", month_columns)
        month_input = input("Which months would you like to keep? (Enter comma-separated months, e.g., Jan, Feb, Mar): ").strip()
        
        # Split the input and clean each month
        months_to_keep = [month.strip().title() for month in month_input.split(',')]
        
        # Validate all months
        invalid_months = [month for month in months_to_keep if month not in month_columns]
        if invalid_months:
            raise ValueError(f"Invalid month(s): {invalid_months}. Please choose from {month_columns}")
        
        print(f"\nUsing month columns: {months_to_keep}")
        
        # Keep only specified columns
        base_columns = [
            'PMT Title', 'Cost Center', 'Application',
            'Resource ATTUID'
        ] + months_to_keep
        filtered_data = filtered_data[base_columns]
        
        # Save to new Excel file
        output_file = os.path.join(script_dir, 'cmpm_modified.xlsx')
        
        # First save the file with basic data
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Write first 8 rows exactly as they are
            first_rows.to_excel(writer, sheet_name='Modified Data', index=False, header=False)
            
            # Create modified header row with new columns
            header_indices = [header_row.index(col) for col in base_columns]
            modified_header = [header_row[i] for i in header_indices] + ['Rate'] + [f'{month} TC' for month in months_to_keep] + ['Location', 'Employee']
            
            # Write the modified header row (9th row)
            pd.DataFrame([modified_header]).to_excel(
                writer, sheet_name='Modified Data',
                startrow=8, index=False, header=False
            )
            
            # Write the filtered data
            filtered_data.to_excel(
                writer, sheet_name='Modified Data',
                startrow=9, index=False, header=False
            )
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Modified Data']
            
            # Calculate column positions
            base_col_count = len(base_columns)
            rate_col = openpyxl.utils.get_column_letter(base_col_count + 1)
            
            # Add Rate column header
            worksheet[f'{rate_col}9'] = 'Rate'
            
            # Add individual total cost columns for each month
            current_col = base_col_count + 2
            total_cost_cols = {}
            
            for month in months_to_keep:
                col_letter = openpyxl.utils.get_column_letter(current_col)
                total_cost_cols[month] = col_letter
                worksheet[f'{col_letter}9'] = f'{month} TC'  # Header for total cost column
                month_col = openpyxl.utils.get_column_letter(base_columns.index(month) + 1)
                
                # Add formulas for Total Cost for each month
                for row in range(10, 10 + len(filtered_data)):
                    worksheet[f'{col_letter}{row}'] = f'={month_col}{row}*{rate_col}{row}'
                
                current_col += 1
            
            # Add subtotal row
            last_row = 9 + len(filtered_data)
            subtotal_row = last_row + 1
            
            # Add subtotals for each month's hours and total cost without the "Subtotal" label
            worksheet[f'A{subtotal_row}'] = ''  # Empty cell instead of "Subtotal"
            
            # Add subtotals for each month's hours and total cost
            for month in months_to_keep:
                # Subtotal for month hours
                month_col = openpyxl.utils.get_column_letter(base_columns.index(month) + 1)
                worksheet[f'{month_col}{subtotal_row}'] = f'=SUBTOTAL(9,{month_col}10:{month_col}{last_row})'
                
                # Subtotal for month's total cost
                tc_col = total_cost_cols[month]
                worksheet[f'{tc_col}{subtotal_row}'] = f'=SUBTOTAL(9,{tc_col}10:{tc_col}{last_row})'
            
            # Add borders to Modified Data sheet
            last_col = openpyxl.utils.get_column_letter(len(modified_header))
            data_range = worksheet[f'A9:{last_col}{subtotal_row}']
            
            # Add thin borders to all cells in the range
            for row in data_range:
                for cell in row:
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )
            
            # Add thick border to subtotal row
            subtotal_range = worksheet[f'A{subtotal_row}:{last_col}{subtotal_row}']
            for cell in subtotal_range[0]:
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thick'),
                    bottom=openpyxl.styles.Side(style='thick')
                )
            
            # Add thick outside border to the entire table
            data_range = worksheet[f'A9:{last_col}{subtotal_row}']
            for row in data_range:
                for cell in row:
                    # Get current cell borders
                    current_border = cell.border
                    
                    # Determine if cell is on an edge
                    is_left_edge = cell.column == 1  # Column A
                    is_right_edge = cell.column == len(modified_header)
                    is_top_edge = cell.row == 9
                    is_bottom_edge = cell.row == subtotal_row
                    
                    # Create new border keeping internal borders thin and making outer borders thick
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thick' if is_left_edge else 'thin'),
                        right=openpyxl.styles.Side(style='thick' if is_right_edge else 'thin'),
                        top=openpyxl.styles.Side(style='thick' if is_top_edge else current_border.top.style),
                        bottom=openpyxl.styles.Side(style='thick' if is_bottom_edge else current_border.bottom.style)
                    )
            
            # Top border
            for col in range(1, len(modified_header) + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                worksheet[f'{col_letter}9'].border = openpyxl.styles.Border(
                    left=worksheet[f'{col_letter}9'].border.left,
                    right=worksheet[f'{col_letter}9'].border.right,
                    top=openpyxl.styles.Side(style='thick'),
                    bottom=worksheet[f'{col_letter}9'].border.bottom
                )
            
            # Bottom border
            for col in range(1, len(modified_header) + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                worksheet[f'{col_letter}{subtotal_row}'].border = openpyxl.styles.Border(
                    left=worksheet[f'{col_letter}{subtotal_row}'].border.left,
                    right=worksheet[f'{col_letter}{subtotal_row}'].border.right,
                    top=worksheet[f'{col_letter}{subtotal_row}'].border.top,
                    bottom=openpyxl.styles.Side(style='thick')
                )

        # Wait a moment to ensure file is released
        time.sleep(1)
        
        # Now create pivot table using Excel automation
        excel = win32.Dispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # Open the workbook
        wb = excel.Workbooks.Open(os.path.abspath(output_file))
        
        # Select Modified Data sheet
        ws = wb.Worksheets("Modified Data")
        
        # Define the data range including headers
        last_row = ws.Cells(ws.Rows.Count, "A").End(-4162).Row
        last_col = openpyxl.utils.get_column_letter(len(modified_header))
        
        # Important: Start from row 9 to include only the actual header and data
        data_range = ws.Range(f"A9:{last_col}{last_row}")
        
        # Create pivot cache
        ws = wb.Worksheets("Modified Data")
        data_range = ws.Range(f"A9:{last_col}{last_row}")
        pc = wb.PivotCaches().Create(
            SourceType=1,  # xlDatabase = 1
            SourceData=data_range,
            Version=6
        )
        
        try:
            # Create new sheet for pivot table
            pt_sheet = wb.Worksheets.Add()
            pt_sheet.Name = "Pivot Table"
            
            # Create pivot table
            pt = pc.CreatePivotTable(
                TableDestination=pt_sheet.Range("A3"),
                TableName="PivotTable1"
            )
            
            # Debug print available fields
            print("\nAvailable fields in pivot table:")
            for field in pt.PivotFields():
                print(f"- {field.Name}")
            
            # Add fields to pivot table
            pt.PivotFields("Cost Center").Orientation = 1  # xlRowField = 1
            pt.PivotFields("Location").Orientation = 1     # xlRowField = 1
            pt.PivotFields("Employee").Orientation = 1     # xlRowField = 1
            
            # Add only the total cost columns for each month
            for month in months_to_keep:
                tc_field_name = f"{month} TC"
                print(f"Adding field: {tc_field_name}")
                tc_field = pt.PivotFields(tc_field_name)
                tc_field.Orientation = 4      # xlDataField = 4
                tc_field.Function = -4157     # xlSum = -4157
            
            # Refresh the pivot table
            pt.RefreshTable()
            
            # Add borders to Pivot Table sheet pivot table
            pt_range = pt.TableRange2
            
            # Add borders to all cells
            pt_range.Borders.LineStyle = 1  # xlContinuous = 1
            pt_range.Borders.Weight = 2     # xlThin = 2
            
            # Add thick outside border
            pt_range.BorderAround(LineStyle=1, Weight=4)  # xlThick = 4
                
        except Exception as e:
            print(f"Error creating pivot table: {str(e)}")
            # Print available fields for debugging
            try:
                print("\nAvailable fields in pivot table:")
                for field in pt.PivotFields():
                    print(f"- {field.Name}")
            except:
                pass
            raise

        # Save and close properly
        wb.Save()
        wb.Close(SaveChanges=True)
        excel.Quit()

        print(f"\nModified file saved as: cmpm_modified.xlsx")
        
    except Exception as e:
        print(f"An error occurred: {str(e)}")
    finally:
        # Make sure Excel is properly closed
        try:
            if wb is not None:
                wb.Close(SaveChanges=True)
            if excel is not None:
                excel.Quit()
        except:
            pass
        
        # Wait a moment to ensure Excel is fully closed
        time.sleep(1)

if __name__ == "__main__":
    modify_excel_file()
