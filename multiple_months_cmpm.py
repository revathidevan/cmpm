import pandas as pd
import os
import openpyxl
from datetime import datetime
import win32com.client as win32
import time

def modify_excel_file():
    excel = None
    wb = None
    try:
        # Get the directory where the script is located
        script_dir = os.path.dirname(os.path.abspath(__file__))
        input_file = os.path.join(script_dir, 'cmpm.xlsx')
        
        # Read the entire Excel file
        df_full = pd.read_excel(input_file, header=None)
        
        # Keep first 8 rows exactly as they are (rows 1-7 and empty row 8)
        first_rows = df_full.iloc[:8].copy()
        
        # Get the header row (9th row)
        header_row = df_full.iloc[8].tolist()  # Convert to list
        
        # Get the data rows (after 9th row)
        data_rows = df_full.iloc[9:].copy()
        data_rows.columns = header_row  # Set the 9th row as header for data
        
        # Filter Cost Center to EY1Z18000
        filtered_data = data_rows[data_rows['Cost Center'] == 'EY1Z18000']
        
        # Get list of month columns
        month_columns = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                        'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        
        # Show available months and ask user
        print("\nAvailable months:", month_columns)
        month_input = input("Which month would you like to keep? (e.g., Jan, Feb, etc.): ").strip()
        
        # Convert input to title case (first letter capital, rest lowercase) for consistency
        month_to_keep = month_input.title()
        
        if month_to_keep not in month_columns:
            raise ValueError(f"Invalid month. Please choose from {month_columns}")
        
        print(f"\nUsing month column: {month_to_keep}")
        
        # Keep only specified columns
        base_columns = [
            'PMT Title', 'Cost Center', 'Application',
            'Resource ATTUID',
            month_to_keep
        ]
        filtered_data = filtered_data[base_columns]
        
        # Save to new Excel file
        output_file = os.path.join(script_dir, 'cmpm_modified.xlsx')
        
        # First save the file with basic data
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Write first 8 rows exactly as they are
            first_rows.to_excel(writer, sheet_name='Modified Data', index=False, header=False)
            
            # Create modified header row with new columns
            header_indices = [header_row.index(col) for col in base_columns]
            modified_header = [header_row[i] for i in header_indices] + ['Rate', 'Total Cost', 'Location', 'Employee']
            
            # Write the modified header row (9th row)
            pd.DataFrame([modified_header]).to_excel(
                writer, sheet_name='Modified Data',
                startrow=8, index=False, header=False
            )
            
            # Write the filtered data
            filtered_data.to_excel(
                writer, sheet_name='Modified Data',
                startrow=9, index=False, header=False
            )
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Modified Data']
            
            # Calculate column positions and add formulas
            base_col_count = len(base_columns)
            month_col = openpyxl.utils.get_column_letter(base_columns.index(month_to_keep) + 1)
            rate_col = openpyxl.utils.get_column_letter(base_col_count + 1)
            total_cost_col = openpyxl.utils.get_column_letter(base_col_count + 2)
            
            # Define last_col here
            last_col = openpyxl.utils.get_column_letter(len(modified_header))
            
            # Add formulas for Total Cost
            for row in range(10, 10 + len(filtered_data)):
                worksheet[f'{total_cost_col}{row}'] = f'={month_col}{row}*{rate_col}{row}'
            
            # Add subtotal row
            last_row = 9 + len(filtered_data)
            subtotal_row = last_row + 1
            
            # Add "Subtotal" label and formulas
            worksheet[f'A{subtotal_row}'] = 'Subtotal'
            worksheet[f'{month_col}{subtotal_row}'] = f'=SUBTOTAL(9,{month_col}10:{month_col}{last_row})'
            worksheet[f'{total_cost_col}{subtotal_row}'] = f'=SUBTOTAL(9,{total_cost_col}10:{total_cost_col}{last_row})'

            # Add borders to Modified Data sheet
            data_range = worksheet[f'A9:{last_col}{subtotal_row}']
            
            # Add borders to all cells in the range
            for row in data_range:
                for cell in row:
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )
            
            # Add thick border around the entire table
            for row in range(9, subtotal_row + 1):
                # Left border
                worksheet[f'A{row}'].border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thick'),
                    right=worksheet[f'A{row}'].border.right,
                    top=worksheet[f'A{row}'].border.top,
                    bottom=worksheet[f'A{row}'].border.bottom
                )
                # Right border
                worksheet[f'{last_col}{row}'].border = openpyxl.styles.Border(
                    left=worksheet[f'{last_col}{row}'].border.left,
                    right=openpyxl.styles.Side(style='thick'),
                    top=worksheet[f'{last_col}{row}'].border.top,
                    bottom=worksheet[f'{last_col}{row}'].border.bottom
                )
            
            # Top border
            for col in range(1, len(modified_header) + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                worksheet[f'{col_letter}9'].border = openpyxl.styles.Border(
                    left=worksheet[f'{col_letter}9'].border.left,
                    right=worksheet[f'{col_letter}9'].border.right,
                    top=openpyxl.styles.Side(style='thick'),
                    bottom=worksheet[f'{col_letter}9'].border.bottom
                )
            
            # Bottom border
            for col in range(1, len(modified_header) + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                worksheet[f'{col_letter}{subtotal_row}'].border = openpyxl.styles.Border(
                    left=worksheet[f'{col_letter}{subtotal_row}'].border.left,
                    right=worksheet[f'{col_letter}{subtotal_row}'].border.right,
                    top=worksheet[f'{col_letter}{subtotal_row}'].border.top,
                    bottom=openpyxl.styles.Side(style='thick')
                )

        # Wait a moment to ensure file is released
        time.sleep(1)
        
        # Now create pivot table using Excel automation
        excel = win32.Dispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # Open the workbook
        wb = excel.Workbooks.Open(os.path.abspath(output_file))
        
        # Select Modified Data sheet
        ws = wb.Worksheets("Modified Data")
        
        # Define the data range including headers
        last_row = ws.Cells(ws.Rows.Count, "A").End(-4162).Row
        last_col = openpyxl.utils.get_column_letter(len(modified_header))
        
        # Important: Start from row 9 to include only the actual header and data
        data_range = ws.Range(f"A9:{last_col}{last_row}")
        
        # Create pivot cache with explicit header row
        pc = wb.PivotCaches().Create(
            SourceType=1,
            SourceData=data_range,
            Version=6
        )
        
        # Create new sheet for pivot table
        pt_sheet = wb.Worksheets.Add()
        pt_sheet.Name = "Summary"
        
        # Create pivot table
        pt = pc.CreatePivotTable(
            TableDestination="Summary!R3C1",
            TableName="SummaryPivotTable", 
            DefaultVersion=6
        )
        
        # Debug: Print the actual field names
        print("\nAvailable fields in pivot table:")
        for field in pt.PivotFields():
            print(f"- {field.Name}")
        
        try:
            # Add Application as filter
            pt.PivotFields("Application").Orientation = 3  # xlPageField (Filter)
            
            # Add Location and Employee as rows
            pt.PivotFields("Location").Orientation = 1     # xlRowField
            pt.PivotFields("Employee").Orientation = 1     # xlRowField
            
            # Add the month and Total Cost as values
            values_field = pt.PivotFields(month_to_keep)
            values_field.Orientation = 4  # xlDataField (Values)
            values_field.Function = -4157  # xlSum
            
            total_cost_field = pt.PivotFields("Total Cost")
            total_cost_field.Orientation = 4  # xlDataField (Values)
            total_cost_field.Function = -4157  # xlSum
            
        except Exception as e:
            print(f"\nError while configuring pivot fields: {str(e)}")
            # Print available fields for debugging
            print("\nAvailable fields in pivot table:")
            for field in pt.PivotFields():
                print(f"- {field.Name}")
            raise

        # Refresh the pivot table
        pt.RefreshTable()
        
        # Add borders to Summary sheet pivot table
        try:
            pt_sheet = wb.Worksheets("Summary")
            pt_range = pt.TableRange2
            
            # Add borders to all cells
            pt_range.Borders.LineStyle = 1  # xlContinuous
            pt_range.Borders.Weight = 2     # xlThin
            
            # Add thick outside borders
            pt_range.BorderAround(LineStyle=1, Weight=4)  # Weight=4 for thick border
            
        except Exception as e:
            print(f"Error adding borders to pivot table: {str(e)}")

        # Save and close properly
        wb.Save()
        wb.Close(SaveChanges=True)
        excel.Quit()

        print(f"\nModified file saved as: cmpm_modified.xlsx")
        
    except Exception as e:
        print(f"An error occurred: {str(e)}")
    finally:
        # Make sure Excel is properly closed
        try:
            if wb is not None:
                wb.Close(SaveChanges=True)
            if excel is not None:
                excel.Quit()
        except:
            pass
        
        # Wait a moment to ensure Excel is fully closed
        time.sleep(1)

if __name__ == "__main__":
    modify_excel_file() 
